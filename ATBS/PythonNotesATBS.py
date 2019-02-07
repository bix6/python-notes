#Automate The Boring Stuff Notes

#Ch1
#operators
2 ** 3 # 2^3
22 % 8 #6
22//8 #2 Integer division/floored quotient
22/8 #2.75
3*5 #15
5 - 2 #3
2+2 #4

#Strings
'Jack'+'Jack'
'Jack'*5

#variable declaration
myNum= 66

#casting
str(0)
int('42')
float(10)

#rounding
int(7.7) # 7
int(7.7) + 1 #8

# equivalence
42 == '42' #False
42 == 42.0 #True
42.0 == 0042.000 #True


#Ch2 - Flow Control
#Booleans
#True and False

#Comparison Operators
# ==, !=, <, >, <=, >=

# Boolean Operators - order of ops = not -> and -> or
True and True #True
True and False #False
False and True # False
False and False # False

True or True #True
True or False #True
False or True #True
False or False #False

not True #False
not not True #True
not False #True

# Mix of Comparison and Boolean Operators
(4<5) and (5<6) #True
2+2==4 and not 2+2==5 and 2*2==2+2 #True

# Conditions (evaluate to Bool) and clauses
name = 'Mary'
password = 'swordfish'
if name == 'Mary':
    print('Hello Mary')
    if password == 'swordfish':
         print('Access granted.')
    else:
        print('Wrong password.')

name = 'Jack'
if name == 'Bob':
    print('Hello B')
elif name == 'Jack': # elif called if all prior conditions are False
    print('Hello J')
elif name == 'Mary':
    print('Hello M') # won't reach here since prior condition is True
else:
    print('Fail') # else guarantees at least one clause will execute. Elif does not.

spam = 0
while spam < 5:
    print('spam: ' + str(spam))
    spam = spam + 1

# Exit infinite loop with Ctrl-C
# 0, 0.0 and '' are all False

#yourName.py progs
name = ''
while name != 'Jack':
    print('Please type your name:')
    name = input()
print('Thank you!')

while True:
    print('Please type your name:')
    name = input()
    if name == 'JackJack':
        break
print('Thank you!')

while True:
    print('Who are you?')
    name = input()
    if name != 'Joe':
        continue
    print('What is the password, Joe?')
    password = input()
    if password == 'swordfish': # This needs another loop cuz it restarts loop if password is wrong
        break
print('Access granted.')

name = ''
while not name:
    print('Enter your name:')
    name = input()
print('How many guests will you have?')
numOfGuests = int(input())
if numOfGuests:
    print('Be sure to have enough room for all your guests.')
print('Done')


print('My name is')
for i in range(5): 
    print('Jimmy Five Times (' + str(i) + ')') # can use break and continue statements here
    print('i:' + str(i))

for i in range(12, 16): #[start, stop)
    print(i)

for i in range(0, 10, 2): #start, stop, step
    print(i)

for i in range(5, -1, -1):
    print(i)


import random
for i in range(5):
    print(random.randint(1,10)) #1-10 inclusive


#Chapter 3
#Create a function
def hello(name):
    print('Howdy!')
    print('AHHHHH!!!! It\'s ' + name)

hello('Bobby')
hello('Jerry')

#magic8Ball.py

# in python None == null
# python adds return None to any function with no return statement (e.g. print())
# just like how while and for loops implicitly end with continues
# if you just type return then None is returned

print('Hello', end='') #disable new line
print('World')

print('cats', 'dogs', 'mice') #cats dogs mice
print('cats', 'dogs', 'mice', sep=',') #cats,dogs,mice

# local variables can use global vars. global cannot use local. local cannot use other local.
# will use local var if global and local var have same name but avoid doing this


def spam():
    global eggs #global tells func to use global var
    eggs = 'spam' #sets global eggs var to 'spam'

def bacon():
    eggs='bacon' #local

def ham():
    print(eggs) #global
    
eggs = 'global'
print(eggs) #'global'
spam()
print(eggs) #'spam'
bacon()
print(eggs) #'spam'
ham() #prints 'spam'


#try and except (try catch)
def spam(divideBy):
    try:
        return 42 / divideBy
    except ZeroDivisionError:
        print('Error: Invalid argument.')

print(spam(2))
print(spam(12))
print(spam(0))
print(spam(1))


#Chapter 4 Lists

# list contains multiple values in ordered sequence
# list value == the list itself (a val that can be stored in a var or passed to function)
# list val looks like ['cat', 'bat', 'rat', 'elephant']
# values inside list are called items and are comma-delimited 
# value[] is empty list similar to ''

spam = ['cat', 'bat', 'rat', 'elephant']
print(spam[0]) #cat
print(spam[1]) #bat
print(spam[-1]) # last = elephant
print(spam[-3]) # 3rd from last = bat
print(spam[0:4]) # slice entire list ['cat', 'bat', 'rat', 'elephant']
print(spam[1:3]) # ['bat', 'rat']
print(spam[0:-1]) # ['cat', 'bat', 'rat]
print(spam[:2]) # ['cat', 'bat']
print(spam[:]) # ['cat', 'bat', 'rat', 'elephant'] entire list
spam[1] = 'aardvark'
print(spam)
print([1,2,3]+['A','B','C']) # add lists
print([1,2,3]*3) # multiply list
del spam[2] # delete item
print(spam)

spam = [['cat', 'bat'], [10, 20, 30, 40, 50]]
print(spam[0]) #['cat', 'bat']
print(spam[0][0]) #cat
print(spam[1][4]) #50
print(len(spam)) # spam length = 2


for i in [0,2,4,6]:
    print(i)

# he refers to sequences as "list-like"

supplies = ['pens', 'staplers', 'flame-throwers', 'binders']
for i in range(len(supplies)): # loop through entier list
    print('Index ' + str(i) + ' in supplies is: ' + supplies[i])

'pens' in supplies # true
'pens' not in supplies # false

cat = ['fat', 'orange', 'loud']
size, color, disposition = cat #sets vars size = fat, color = orange, disposition = loud

a, b = 'Alice', 'Bob'
a, b = b, ca #swaps Alice and Bob -- typo here?

spam = ['hello', 'hi', 'howdy', 'heyas']
spam.index('hello') # 0, only returns 1st instance
spam.append('moose')
spam.insert(1, 'chicken') #inserts after?
spam.remove('chicken')
spam.sort()
spam.sort(reverse=True)
spam.sort(key=str.lower)

print('Four score and seven ' + \
      'years ago...') # line continuation

name = 'Zophie'
name[0] # 'Z'
name[-2] # 'i'
name[0:4] # 'Zoph'
'Zo' in name # True
for i in name:
    print('***' + i + '***')

name= 'Zophie a cat'
newName=name[0:7]+'the'+name[8:12]

eggs = ('hello', 42, 0.5) # tuple uses () and is immutable

tuple(['cat', 'dog', 5]) # convert lists and tuples into eachother
list(('cat', 'dog', 5))
list('hello')

import copy
cheese = copy.copy(spam) # creates a copy of a tuple/list. .deepcopy() if list has lists



#Chapter 5 - dictionaries and data structuring
# indexes for dictionaries are called keys and they have a value. key-value pair
myCat = {'size': 'fat', 'color':'gray', 'disposition':'loud'}
myCat['size']
spam = {12345:'Luggage Combo', 42:'The Answer'} # using integers as keys

birthdays = {'Alice':'Apr 1', 'Bob':'Dec 12', 'Carol':'Mar 4'}

while True:
    print('Enter a name: (blank to quit)')
    name = input()
    if name == '':
        break

    if name in birthdays:
        print(birthdays[name] + ' is the birthday of ' + name)
    else:
        print('I do not have birthday information for ' + name)
        print('What is their birthday?')
        bday = input()
        birthdays[name] = bday
        print('Birthday database updated.')


spam = {'color':'red', 'age':42}
for k in spam.keys(): #prints key vals
    print(k)
for v in spam.values(): # prints value vals
    print(v)
for i in spam.items(): #prints tuples of key-value pairs
    print(i)
print(list(spam.keys())) # make a list of keys
for k,v in spam.items(): #multiple assignment to print both keys and values
    print('Key: ' + k + ' Value: ' + str(v))
'color' in spam.keys()
'color' in spam # shorthand of spam.keys()
'red' not in spam.values()

picnicItems = {'apples':5, 'cups':2}
print('I am bringing ' + str(picnicItems.get('cups', 0)) + ' cups.') # get returns value if it exits 
print('I am bringing ' + str(picnicItems.get('eggs', 0)) + ' cups.') # otherwise prints fall back
spam.setdefault('color','black') # since color already exists will return 'red'
spam.setdefault('flavor','yummy') # creates a key-val pair since 'type' doesn't exist yet


import pprint
message = 'It was a bright cold day in April, and the clocks were striking thirteen.'
count = {}
for character in message:
    count.setdefault(character, 0)
    count[character] = count[character] + 1
pprint.pprint(count) # pretty pring dict vals
myStr = pprint.pformat(count) # pformat gives you the vals as a string

    
allGuests = {'Alice':{'apples':5, 'pretzels':12},
             'Bob':{'ham sandwiches':3,'apples':2},
             'Carol':{'cups':3,'apple pies':1}}
def totalBrought(guests, item):
    numBrought = 0
    for k, v in guests.items():
        numBrought += v.get(item, 0)
    return numBrought
print('Number of things being brought:')
print(' - Apples         ' + str(totalBrought(allGuests, 'apples')))
print(' - Cups           ' + str(totalBrought(allGuests, 'cups')))
print(' - Cakes          ' + str(totalBrought(allGuests, 'cakes')))
print(' - Ham Sandwiches ' + str(totalBrought(allGuests, 'ham sandwiches')))
print(' - Apple Pies     ' + str(totalBrought(allGuests, 'apple pies')))


#Chapter 6 - Strings
print("This is also a string. Jack's string") # can use " or ' for string
print('Say hi to Jack\'s pitbull.') # \ for escape character
# \' \" \t \n \\
print('Hello\nHello Baby')
print(r'That is Jack\'s cat.') # r makes raw string, ignores escape char
print('''Dear Jack,

You smell. Jack's nice nice.

Best,
Jack''') # multline string uses triple '''
'''Multline
comment
wow trippy''' #multline comment
spam = 'Hello world!'
fizz = spam[0:5]
print(fizz)
print('Hello' in 'Hello World') #true
print('' in 'spam') #true
spam=spam.upper()
print(spam)
spam=spam.lower()
print(spam)
print('How are you?')
feeling = input()
if feeling.lower() == 'great': # compare text with .lower and .upper
    print('I feel great too!')
else:
    print('I hope the rest of your day is very nice.')
#isupper() and islower() return True if string has at least 1 letter and all letters are upper or lowercase
print('HELLO'.isupper()) #true
print('abc12345'.islower()) #true
print('ABC12345'.isupper()) #true
print('Hello'.upper().lower().upper()) #chaining
print('alphabet'.isalpha()) # .isalpha() True if only letters and not blank
print('alnum123'.isalnum()) # .isalnum() True if only letters and numbers and not blank
print('123'.isdecimal()) #.isdecimal() True if only 'numeric characters' and not blank
print(' \t\n'.isspace()) #.isspace() True if only space, tabs and new lines and not blank
print('Mr Jack'.istitle()) #istitle() True if all words start with uppercase followed by lowercase
print('Hello world!'.startswith('Hello')) #.startswith()
print('Hello world!'.endswith('world!')) #.endswith()
print(', '.join(['cats','rats','bats'])) # .join([]) joins a list into a string
print('ABC'.join(['Easy','as','123'])) 
print('My name is Jack'.split()) #.split() splits a string into a list
print('MyABCnameABCisABCJack'.split('ABC')) 
print('My\nname\nis\nJack'.split('\n')) #often used to split newlines
print('Hello'.rjust(10)) #right justifies with 10 total spaces (5 for Hello and 5 spaces)
print('Hello'.ljust(20)) #left justifies with 20 total spaces (5 for Hello and 15 spaces)
print('Hello'.ljust(20, '*')) #fill char
print('Hello'.center(20, '-')) #.center()

def printPicnic(itemsDict, leftWidth, rightWidth):
    print('PICNIC ITEMS'.center(leftWidth+rightWidth,'-'))
    for k, v in itemsDict.items():
        print(k.ljust(leftWidth,'.')+str(v).rjust(rightWidth))
picnicItems={'sandwiches':4,'apples':12,'cups':4,'cookies':8000}
printPicnic(picnicItems,12,5)
printPicnic(picnicItems,20,6)

spam = '     Hello World      '
print(spam.strip()) #strips all whitespace (incl tab and newline)
print(spam.lstrip())
print(spam.rstrip()) # left and right
spam = 'SpamSpamBaconSpamEggsSpamSpam'
print(spam.strip('ampS')) #strips Spam from ends, order of letters doesnt matter


#Appendix A
#pip install pyperclip
# make sure to be admin when changing PATH

#Appendix B
#shebang line = first line to tell computer this is python = #! python3.
# USE THIS SHEBANG LINE #!/usr/bin/env python3
# JK just use #! python3 I think

#Create .bat batch file using py.exe
#create text file like this @py.exe C:\Users\JackB\Desktop\Python\collatz.py %*
# and save as collatz.bat
# run using win-R and typing in name



#Chapter 7 - Pattern matching with regular expressions
def isPhoneNumber(text): #bulky way without regular expressions
    if len(text) != 12:
        return False
    for i in range(0,3):
        if not text[i].isdecimal():
            return False
    if text[3] != '-':
        return False
    for i in range(4,7):
        if not text[i].isdecimal():
            return False
    if text[7] != '-':
        return False
    for i in range(8,12):
        if not text[i].isdecimal():
            return False
    return True

print('415-555-4242 is a phone number:')
print(isPhoneNumber('415-555-4242'))
print('Moshi moshi is a phone number:')
print(isPhoneNumber('Moshi moshi'))

message = 'Call me at 415-555-1011 tomorrow. 415-555-9999 is my office.'
for i in range(len(message)):
    chunk = message[i:i+12]
    if isPhoneNumber(chunk):
        print('Phone number found: ' + chunk)

# if we use regular expressions (regexes) we can simplify the code
# \d\d\d-\d\d\d-\d\d\d\d is the equivalent regex for the above phone numbers
#\d{3}-\d{3}-\d{4} is equivalent. {x} says Match this pattern x times
# verify regexes online 
# import re

import re
phoneNumRegex = re.compile(r'\d{3}-\d{3}-\d{4}') #use r to make a raw string
mo = phoneNumRegex.search('My number is 415-555-4242') #.search returns None or a Match object if found
print('Phone number found: ' + mo.group()) #.group gets the Match object

phoneNumRegex = re.compile(r'(\d{3})-(\d{3}-\d{4})') # () make groups 1,2,3,etc.
mo = phoneNumRegex.search('My number is 415-555-4242') #.search returns None or a Match object if found
print('Area Code: ' + mo.group(1)) # first set of ()
print('Partial Number: ' + mo.group(2))
print('Entire Number: ' + mo.group(0)) # .group() and .group(0) are equivalent

areaCode, mainNumber = mo.groups() # . groups() returns a tuple of groups as strings
print(mo.groups()) #('415', '555-4242')
print(areaCode) #415
print(mainNumber)

phoneNumRegex = re.compile(r'(\(\d{3}\)) (\d{3}-\d{4})') # use escape chars \( \)to search for () 
mo = phoneNumRegex.search('My phone number is (415) 555-4242')
print(mo.groups())

# use the pipe | to match one of many (or)
heroRegex = re.compile(r'Batman|Tina Fey') #Batman or Tina Fey, returns 1st instance
mo1 = heroRegex.search('Batman and Tina Fey')
print(mo1.group())
mo2 = heroRegex.search('Tina Fey and Batman')
print(mo2.group())

batRegex = re.compile(r'Bat(man|mobile|copter|bat)') #Batman or Batmobile or Batcopter or Batbat
mo = batRegex.search('Batmobile lost a wheel')
print(mo.group()) #Batmobile
print(mo.group(1)) #mobile

batRegex = re.compile(r'Bat(wo)?man') # ? is an optional part of the pattern
mo1 = batRegex.search('The Adventures of Batman')
print(mo1.group())
mo2 = batRegex.search('The Adventures of Batwoman')
print(mo2.group())

phoneRegex = re.compile(r'(\d{3}-)?\d{3}-\d{4}') # using ? to find numbers with and without area codes 
mo1 = phoneRegex.search('My number is 415-453-3453')
print(mo1.group())
mo2 = phoneRegex.search('My number is 555-2342')
print(mo2.group())

batRegex = re.compile(r'Bat(wo)*man') # group before * can be repeated 0-infinity times
mo1 = batRegex.search('The Adventures of Batman')
print(mo1.group()) #finds Batman
mo2 = batRegex.search('The Adventures of Batwoman')
print(mo2.group()) #finds Batwoman
mo3 = batRegex.search('The Adventures of Batwowowowowowoman')
print(mo3.group()) #finds Batwowowowoman

batRegex = re.compile(r'Bat(wo)+man') # group before + matched if repeated 1-infinity times
mo1 = batRegex.search('The Adventures of Batwoman')
print(mo2.group())

haRegex = re.compile(r'(Ha){3}') #Matches HaHaHa {x} matchex x times group
mo1 = haRegex.search('HaHaHa')
print(mo1.group())

haRegex = re.compile(r'(Ha){,5}') #Matches 0-5 Ha {,5}
mo1 = haRegex.search('')
print(mo1.group())
mo2 = haRegex.search('Ha')
print(mo2.group())
mo3 = haRegex.search('HaHaHaHaHa') # reg expressions are greedy by default so will match the longest string possible
print(mo3.group())
nongreedyHaRegex = re.compile(r'(Ha){3,5}?') #{3,5}? makes it nongreedy so returns 3 Ha
mo1 = nongreedyHaRegex.search('HaHaHaHaHa') 
print(mo1.group())

phoneNumRegex = re.compile(r'\d{3}-\d{3}-\d{4}') # no groups if using findall
print(phoneNumRegex.findall('Cell: 415-666-6666 Work: 232-239-2983')) #.findall() returns a list
phoneNumRegex = re.compile(r'(\d{3})-(\d{3})-(\d{4})') # can use groups with findall
print(phoneNumRegex.findall('Cell: 415-666-6666 Work: 232-239-2983')) # returns a list of tuples

# Character classes
# \d is shorthand for (0|1|2|3|4|5|6|7|8\9)
# \D any char that is not a numeric digit from 0-9
# \w Any letter, numeric digit or underscore (matching "word" characters)
# \W Any char that is not a letter, numeric digit or the underscore char
# \s any space, tab or newline char
# \S any char but space, tab or newline
# Can create own Char classes using []

xmasRegex = re.compile(r'\d+\s\w+') # find 1 or more digits followed by a space followed by 1 or more chars
print(xmasRegex.findall('12 drummers, 11 pipers, 10 lords, 9 ladies, 8 maids, 7 swans'))

vowelRegex = re.compile(r'[aeiouAEIOU]') #matches any vowels individually
print(vowelRegex.findall('Robocop eats baby food. BABY FOOD.'))
consonantRegex = re.compile(r'[^aeiouAEIOU]') # ^ makes a negative character class (matches anything not in [])
print(consonantRegex.findall('Robocop eats baby food. BABY FOOD.'))
#don't need to escape characters. for example [0-5.] will match 0-5 and a period. don't need \. inside []

beginsWithHello = re.compile(r'^Hello') #^ says matched string must start with Hello
print(beginsWithHello.search('Hello world!'))
print(beginsWithHello.search('He said hello.') == None)

endsWithNumber = re.compile(r'\d$') # match must end with a numeric char
print(endsWithNumber.search('Your number is 42'))
print(endsWithNumber.search('You like 42 dogs') == None)

wholeStringIsNum = re.compile(r'^\d+$') # entire string must be numeric
print(wholeStringIsNum.search('28934579234'))
print(wholeStringIsNum.search('12980sxoi290234') == None)

atRegex = re.compile(r'.at') # . is wildcard so will match any char except newline
print(atRegex.findall('The cat in the hat sat on the flat mat.'))

nameRegex = re.compile(r'First Name: (.*) Last Name: (.*)') # .* finds anything b/c . any char, * zero or more chars
mo = nameRegex.search('First Name: Jack Last Name: Bixby') # greedy so will match as much text as possible
print(mo.group())
print(mo.group(1))
print(mo.group(2))

nongreedyRegex = re.compile(r'<.*?>') # ? makes it non greedy
mo = nongreedyRegex.search('<To serve man> for dinner.>')
print(mo.group())
greedyRegex = re.compile(r'<.*>') # no ? so greedy
mo = greedyRegex.search('<To serve man> for dinner.>')
print(mo.group())

noNewlineRegex = re.compile('.*')
print(noNewlineRegex.search('Serve the public trust.\nProtec the innocent.\nUphold the law.').group())
newlineRegex = re.compile('.*', re.DOTALL) #.DOTALL will match all chars including new line
print(newlineRegex.search('Serve the public trust.\nProtect the innocent.\nUphold the law.').group())

#regex symbol review
# ? matches 0 or 1 of preceding group
# * matches 0 or more of preceding group
# + matches 1 or more of preceding group
# {n} matches exactly n of preceding group
# {n,} matches n or more of preceding group
# {,m} matches 0 to m of preceding group
# {n,m} matches at least n and at most m of preceding group
# {n,m}? or *? or +? performs a nongreedy match of preceding group
# ^spam means the string must begin with spam
# spam$ means the string must end with spam
# . matches any character except newline
# \d, \w and \s match a digit, word or space char respectively
# \D, \W and \S match anything except a digit, word or space char respectively
# [abc] matches any char between brackets (a, b, c)
# [^abc] matches any char that isnt between brackets (d, e, f, etc.)

robocop = re.compile(r'robocop', re.I) #re.I or re.IGNORECASE will ignore case
print(robocop.search('RoBOcoP is part man, part machine, all cop.').group())

namesRegex = re.compile(r'Agent \w+') # .sub('replace with', 'original string')
newStr = namesRegex.sub('CENSORED', 'Agent Alice gave the secret documents to Agent Bob.')
print(newStr) # replaces Agent Alice and Agent Bob with Censored

agentNamesRegex = re.compile(r'Agent (\w)\w*')
newStr = agentNamesRegex.sub(r'\1****', 'Agent Alice told Agent Carol that Agent Eve knew Agent Bob was a double agent.')
print(newStr) # .sub('\1', string) the \1 string will be replaced by whatever group 1 (\w) was found

phoneRegex = re.compile(r'((\d{3}|\(\d{3}\))?(\s|-|\.)?\d{3}(\s|-|\.)\d{4}(\s*(ext|x|ext.)\s*\d{2,5})?)')
phoneRegex = re.compile(r'''(
    (\d{3}|\(\d{3}\))?              # area code
    (\s|-|\.)?                      # separator
    \d{3}                           # first 3 digits
    (\s|-|\.)                       # separator
    \d{4}                           # last 4 digits
    (\s*(ext|x|ext.)\s*\d{2,5})?    #extension
    )''', re.VERBOSE) #Verbose ignores whitespace and comments to aid formatting

someRegexValue = re.compile('food', re.IGNORECASE | re.DOTALL | re.VERBOSE) # use | to use multiple second args




# Chapter 8 - Reading and Writing Files
# On Windows root is C:\
# On Mac/Linux root is /

import os
print(os.path.join('usr', 'bin', 'spam')) # returns correct separators depending on OS
#returns usr\\bin\\spam then prints usr\bin\spam

myFiles = ['accounts.txt', 'details.csv', 'invite.docx']
for filename in myFiles:
    print(os.path.join('C:\\Users\\JBL', filename))
# prints C:\Users\JBL\accounts.txt, etc.

print(os.getcwd()) # prints current working directory
# os.chdir('C:\\ThisFolderDoesNotExist')
# os.chidir changes to the specified directory

# Absolute path always starts with root
# Relative path, relative to program's current working directory
# . is short for this directory
# .. means parent folder

# os.makedirs('C:\\delicious\\walnut\\waffles')
# makes entire hierarchy of folders if they don't already exists

print(os.path.abspath('.')) # prints absolute path of cwd
print(os.path.isabs('.')) # returns False since this is not an absolute path
print(os.path.isabs(os.path.abspath('.'))) # True

#relpath(path, start) returns string of relative path from start to path. If no start, uses cwd
print(os.path.relpath('C:\\Windows', 'C:\\')) #Windows
print(os.path.relpath('C:\\Windows', 'C:\\spam\\eggs')) #..\\..\\Windows

#os.path.dirname(path) returns string of everything that comes before last slash in arg
#os.path.basename(path) returns string of everything that comes after last slash in arg
path = 'C:\\Windows\\System32\\calc.exe'
print(os.path.basename(path))
print(os.path.dirname(path))
#os.path.split returns tuple with both dir and basename
print(os.path.split(path))
# os.sep var is set to correct folder separating slash for OS
print(path.split(os.path.sep)) #splits path whenever it finds the proper folder separator, returns as list

#os.path.getsize(path) returns size in bytes of file
#os.listdir(path) will return list of filename strings for each file in arg
print(os.path.getsize('C:\\Windows\\System32\\calc.exe'))
print(os.listdir('C:\\Users\\JackB\\Desktop\\Logitech'))

#find total size of all files in a dir, combine listdir and getsize
totalSize = 0
for filename in os.listdir('C:\\Users\\JackB\\Desktop\\Logitech'):
    totalSize = totalSize + os.path.getsize(os.path.join('C:\\Users\\JackB\\Desktop\\Logitech', filename))
print(totalSize)

# os.path.exists(path) returns True if file or folder exists, False otherwise
# os.path.isfile(path) returns True if path exists and is file, False otherwise
# os.path.isdir(path) returns True if path exists and is folder, False otherwise
print(os.path.exists('C:\\Windows'))
print(os.path.isdir('C:\\Users\\JackB\\Desktop\\Logitech'))
print(os.path.isfile('C:\\Users\\JackB\\Desktop\\Logitech'))
print(os.path.exists('D:\\')) # can use this to check if additional drives are attached, e.g. flash drive

# These functions apply to plaintext files only. Plaintext contain only basic text chars. .txt and .py are
# examples of plaintext. Can treat contents as ordinary string values
# Binary files are all other types such as PDFs and images. Looks like scrambled nonsense in Notepad.
# Every type of Binary file must be handled differently. Many modules exist to make handling them easier.

# 3 steps to read/write files in Python
# 1. Call open() to return a File obj
# 2. Call read() or write() on File obj
# 3. Close file using close() on File obj
# 'C:\\Users\\JackB\\Desktop\\PythonRW\\Hello.txt'

helloFile = open('C:\\Users\\JackB\\Desktop\\Python\\PythonRW\\Hello.txt')
#helloFile = open('C:\\Users\\JackB\\Desktop\\PythonRW\\Hello.txt', 'r')
# 'r' flag makes this read only. These are equivalent since 'r' is default if no value passed

print(helloFile) #prints textwrapper with path/name, mode (r) and encoding

helloContent = helloFile.read()
print(helloContent)

sonnetFile = open('C:\\Users\\JackB\\Desktop\\Python\\PythonRW\\Sonnet29.txt')
listOfLines = sonnetFile.readlines() # read lines returns a list of each line's content
print(listOfLines)

# to write to file need to open in either write mode or append mode
# 'w' = write and overwrites entire file
# 'a' = append and appends changes to end of file
# if filename doesn't exist, both modes will create a new blank file
# after reading or writing, call close() before opening file again

os.chdir('C:\\Users\\JackB\\Desktop\\Python\\PythonRW')
print(os.getcwd())
baconFile = open('bacon.txt', 'w')
baconFile.write('Hello bacon!\n') # write returns the number of chars written, here we do nothing with the return
baconFile.close()
baconFile = open('bacon.txt', 'a')
baconFile.write('Bacon is not a vegetable.')
baconFile.close()
baconFile = open('bacon.txt')
content = baconFile.read()
baconFile.close()
print(content)

# Can save variables to 'binary shelf files' using shelve module
# can restore data to variables from the hard drive
# enables save and open features
# creates 3 new files in CWD - .bak, .dat and .dir (e.g. mydata.bak)
import shelve
shelfFile = shelve.open('mydata')
cats = ['Zophie', 'Pooka', 'Simon']
shelfFile['cats'] = cats # stores list of cats as value to 'cats' key (dictionary)
shelfFile.close()

shelfFile = shelve.open('mydata')
print(type(shelfFile))
print(shelfFile['cats'])
shelfFile.close()

# Can use pprint.pformat() to get pretty print format of list or dictionary. Syntactically correct python code.
# Can use to create my own modules
import pprint
cats = [{'name': 'Zophie', 'desc':'chubby'}, {'name': 'Pooka', 'desc':'fluffy'}]
pprint.pformat(cats)
fileObj = open('myCats.py', 'w')
fileObj.write('cats = ' + pprint.pformat(cats) + '\n')
fileObj.close()

import myCats # I think this only works if myCats is in PATH 
print(myCats.cats)
print(myCats.cats[0])
print(myCats.cats[0]['name'])

# .py file benefit is that it's a text file so anyone can modify
# generally shelving is preferred for variable saving
# only basic data types (int, float, string, list, dict) can be written to file as simple text



# Incomplete Chapter 12 - Excel
import openpyxl
wb = openpyxl.load_workbook('example.xlsx') #returns a Workbook Type
print(type(wb)) #<class 'openpyxl.workbook.workbook.Workbook'>
# wb.get_sheet_names() # This is deprecated, use wb.sheetnames
print(wb.sheetnames) # sheetnames as a list
# sheet1 = wb.get_sheet_by_name('example') #deprecated, use wb[sheetname]
sheet1 = wb['example'] # Gets a worksheet
print(sheet1) # <Worksheet "example">
print(sheet1.title)
activeSheet = wb.active
print(activeSheet)

# Accessing cells
print(sheet1['A1']) #<Cell 'example'.A1>
print(sheet1['A1'].value) #2014-04-05 13:34:00

c = sheet1['B1']
print(c.value)
print('Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value) # cell row and column
print('Cell ' + c.coordinate + ' is ' + c.value) # coordinate

print(sheet1.cell(row=1, column=2)) # rows and columns start at 1 so this is 1B aka B1
print(sheet1.cell(row=1, column=2).value)
for i in range(1,8,2):
    print(i, sheet1.cell(row=i, column=2).value)

print(sheet1.max_row)
print(sheet1.max_column)





